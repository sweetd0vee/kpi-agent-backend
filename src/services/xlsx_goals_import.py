"""Разбор xlsx для таблиц целей (тот же набор заголовков, что и во фронте `importGoals.ts`)."""

from __future__ import annotations

import re
from datetime import date, datetime
from uuid import uuid4
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from src.models.goal_tables import GoalRow
from src.models.leader_goal_tables import LeaderGoalRow
from src.models.strategy_goal_tables import StrategyGoalRow


def _generate_row_id() -> str:
    """UUID v4, в одном формате с `generateId()` на фронте."""
    return str(uuid4())


def _normalize_header_for_match(value: Any) -> str:
    """Как на фронте `xlsxImportHeaders.ts`: пробелы, ё/е, регистр, : и /."""
    s = str(value if value is not None else "")
    s = s.replace("\ufeff", "")
    s = re.sub(r"[\u00a0\u202f\u2007\u2009\u200b\ufeff\t\r\n]+", " ", s)
    s = re.sub(r"\s*/\s*", "/", s)
    s = re.sub(r"\s*:\s*", ":", s)
    s = re.sub(r"\s+", " ", s).strip().lower()
    s = s.replace("ё", "е")
    return s


def _build_header_field_lookup(canonical: dict[str, str]) -> dict[str, str]:
    out: dict[str, str] = {}
    for k, v in canonical.items():
        out[_normalize_header_for_match(k)] = v
    return out


def _normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _read_first_sheet_matrix(content: bytes) -> list[list[Any]]:
    bio = BytesIO(content)
    wb = load_workbook(bio, data_only=True, read_only=True)
    try:
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


BOARD_HEADER_TO_FIELD: dict[str, str] = {
    "ФИО": "lastName",
    "Бизнес/блок": "businessUnit",
    "Бизнес юнит": "businessUnit",
    "Департамент": "department",
    "Подразделение": "department",
    "UUID руководителя": "leaderId",
    "SCAI Цель": "goal",
    "Метрические цели": "metricGoals",
    "вес квартал": "weightQ",
    "Вес квартал": "weightQ",
    "вес год": "weightYear",
    "Вес год": "weightYear",
    "1 квартал": "q1",
    "2 квартал": "q2",
    "3 квартал": "q3",
    "4 квартал": "q4",
    "Отчётный год": "reportYear",
    "Отчетный год": "reportYear",
    "Год": "year",
}

LEADER_HEADER_TO_FIELD: dict[str, str] = {
    "ФИО": "lastName",
    "№ цели": "goalNum",
    "Наименование КПЭ": "name",
    "Тип цели": "goalType",
    "Вид цели": "goalKind",
    "Ед. изм.": "unit",
    "ед.изм.": "unit",
    "Единица измерения": "unit",
    "I кв. Вес %": "q1Weight",
    "I квартал Вес %": "q1Weight",
    "I кв. План. / веха": "q1Value",
    "I квартал Плановое значение": "q1Value",
    "II кв. Вес %": "q2Weight",
    "II квартал Вес %": "q2Weight",
    "II кв. План. / веха": "q2Value",
    "III кв. Вес %": "q3Weight",
    "III кв. План. / веха": "q3Value",
    "IV кв. Вес %": "q4Weight",
    "IV кв. План. / веха": "q4Value",
    "Год Вес %": "yearWeight",
    "Год План. / веха": "yearValue",
    "Комментарии": "comments",
    "Методика расчёта": "methodDesc",
    "Методика расчета": "methodDesc",
    "Источник информации": "sourceInfo",
    "Отчётный год": "reportYear",
    "Отчетный год": "reportYear",
}

STRATEGY_HEADER_TO_FIELD: dict[str, str] = {
    "Бизнес/блок": "businessUnit",
    "Сегмент": "segment",
    "Стратегический приоритет": "strategicPriority",
    "Цель": "goalObjective",
    "Инициатива": "initiative",
    "Тип инициативы": "initiativeType",
    "Ответственный исполнитель": "responsiblePersonOwner",
    "Участие других блоков": "otherUnitsInvolved",
    "Бюджет": "budget",
    "Начало": "startDate",
    "Конец": "endDate",
    "КПЭ": "kpi",
    "ед. изм.": "unitOfMeasure",
    "ед.изм.": "unitOfMeasure",
    "Ед. изм.": "unitOfMeasure",
    "2025: Целевое значение": "targetValue2025",
    "2026: Целевое значение": "targetValue2026",
    "2027: Целевое значение": "targetValue2027",
    "Целевое значение 2025": "targetValue2025",
    "Целевое значение 2026": "targetValue2026",
    "Целевое значение 2027": "targetValue2027",
}


BOARD_HEADER_LOOKUP = _build_header_field_lookup(BOARD_HEADER_TO_FIELD)
LEADER_HEADER_LOOKUP = _build_header_field_lookup(LEADER_HEADER_TO_FIELD)
STRATEGY_HEADER_LOOKUP = _build_header_field_lookup(STRATEGY_HEADER_TO_FIELD)


def parse_board_goals_xlsx(content: bytes) -> list[GoalRow]:
    matrix = _read_first_sheet_matrix(content)
    if len(matrix) < 2:
        return []
    col_to_field: dict[int, str] = {}
    for index, cell in enumerate(matrix[0]):
        field = BOARD_HEADER_LOOKUP.get(_normalize_header_for_match(cell))
        if field:
            col_to_field[index] = field
    if not col_to_field:
        raise ValueError("Не найдено ни одной известной колонки по заголовкам первой строки")

    out: list[GoalRow] = []
    for raw in matrix[1:]:
        row: dict[str, str] = {
            "lastName": "",
            "leaderId": "",
            "businessUnit": "",
            "department": "",
            "goal": "",
            "metricGoals": "",
            "weightQ": "",
            "weightYear": "",
            "q1": "",
            "q2": "",
            "q3": "",
            "q4": "",
            "reportYear": "",
            "year": "",
        }
        for col_index, field in col_to_field.items():
            cell = raw[col_index] if col_index < len(raw) else None
            row[field] = _normalize_cell(cell)
        if not any(str(v).strip() for v in row.values()):
            continue
        leader_raw = (row.get("leaderId") or "").strip()
        lid = leader_raw if leader_raw else None
        out.append(
            GoalRow(
                id=_generate_row_id(),
                lastName=row["lastName"],
                leaderId=lid,
                businessUnit=row["businessUnit"],
                department=row["department"],
                goal=row["goal"],
                metricGoals=row["metricGoals"],
                weightQ=row["weightQ"],
                weightYear=row["weightYear"],
                q1=row["q1"],
                q2=row["q2"],
                q3=row["q3"],
                q4=row["q4"],
                reportYear=row["reportYear"],
                year=row["year"],
            )
        )
    return out


def parse_leader_goals_xlsx(content: bytes) -> list[LeaderGoalRow]:
    matrix = _read_first_sheet_matrix(content)
    if len(matrix) < 2:
        return []
    col_to_field: dict[int, str] = {}
    for index, cell in enumerate(matrix[0]):
        field = LEADER_HEADER_LOOKUP.get(_normalize_header_for_match(cell))
        if field:
            col_to_field[index] = field
    if not col_to_field:
        raise ValueError("Не найдено ни одной известной колонки по заголовкам первой строки")

    empty = {
        "lastName": "",
        "goalNum": "",
        "name": "",
        "goalType": "",
        "goalKind": "",
        "unit": "",
        "q1Weight": "",
        "q1Value": "",
        "q2Weight": "",
        "q2Value": "",
        "q3Weight": "",
        "q3Value": "",
        "q4Weight": "",
        "q4Value": "",
        "yearWeight": "",
        "yearValue": "",
        "comments": "",
        "methodDesc": "",
        "sourceInfo": "",
        "reportYear": "",
    }
    out: list[LeaderGoalRow] = []
    for raw in matrix[1:]:
        row = dict(empty)
        for col_index, field in col_to_field.items():
            cell = raw[col_index] if col_index < len(raw) else None
            row[field] = _normalize_cell(cell)
        if any(row[k] for k in row):
            out.append(LeaderGoalRow(id=_generate_row_id(), **row))
    return out


def parse_strategy_goals_xlsx(content: bytes) -> list[StrategyGoalRow]:
    matrix = _read_first_sheet_matrix(content)
    if len(matrix) < 2:
        return []
    col_to_field: dict[int, str] = {}
    for index, cell in enumerate(matrix[0]):
        field = STRATEGY_HEADER_LOOKUP.get(_normalize_header_for_match(cell))
        if field:
            col_to_field[index] = field
    if not col_to_field:
        raise ValueError("Не найдено ни одной известной колонки по заголовкам первой строки")

    empty = {
        "businessUnit": "",
        "segment": "",
        "strategicPriority": "",
        "goalObjective": "",
        "initiative": "",
        "initiativeType": "",
        "responsiblePersonOwner": "",
        "otherUnitsInvolved": "",
        "budget": "",
        "startDate": "",
        "endDate": "",
        "kpi": "",
        "unitOfMeasure": "",
        "targetValue2025": "",
        "targetValue2026": "",
        "targetValue2027": "",
    }
    out: list[StrategyGoalRow] = []
    for raw in matrix[1:]:
        row = dict(empty)
        for col_index, field in col_to_field.items():
            cell = raw[col_index] if col_index < len(raw) else None
            row[field] = _normalize_cell(cell)
        if any(row.values()):
            out.append(StrategyGoalRow(id=_generate_row_id(), **row))
    return out
